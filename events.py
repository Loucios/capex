import re
from collections import Counter, defaultdict

from openpyxl.utils import get_column_letter


class BaseMixin:

    def __init__(self, events, tso_list) -> None:
        self.events = events
        self.tso_list = tso_list

    def create_main_header(self, ws, tso_name, table):
        row = table.title_row
        column = table.title_column
        ws.cell(row=row, column=column).value = tso_name
        style = table.title_style
        ws.cell(row=row, column=column).style = style
        return ws

    def get_unit_type(self, titles, number):
        unit_type = titles.unit_type
        return self.events[unit_type][number]

    def set_number_format(self,
                          ws,
                          structure,
                          titles,
                          row,
                          column):
        if (structure == titles.mw
                or structure == titles.th):
            ws.cell(row=row, column=column).number_format = '# ##0'

        elif structure == titles.gh:
            ws.cell(row=row, column=column).number_format = '# ##0.00'

        elif (structure == titles.total_cost
              or structure == titles.total
              or structure == titles.capex_flow):
            ws.cell(row=row, column=column).number_format = '# ##0.0'

        return ws

    def get_short_name(self, tso_name):
        if self.tso_list is None:
            pattern = re.compile(r'[^\w ]+')
            return re.sub(pattern, '', tso_name)[:10]
        return self.tso_list[tso_name]

    def get_sum_terms(self, capex_flow, tables, titles):
        '''Create terms like 2024-2026 in last line '''

        first_year = tables.terms[titles.first_year]
        last_year = tables.terms[titles.last_year]
        start_year = 0
        end_year = first_year
        for year in range(last_year - first_year + 1):
            if capex_flow[year]:
                if not start_year:
                    start_year = first_year + year
                end_year = first_year + year
        if not start_year:
            return ''
        if start_year == end_year:
            return start_year
        return f'{start_year} - {end_year}'


class TableMixin:

    def create_table(self,
                     tables,
                     workbook,
                     titles,
                     table):

        # How much events
        events_number = len(self.events[titles.number]) - 1

        # How much events by tso
        tso_frequency = Counter(self.events[titles.tso_name])

        # Count evets by tso
        num_tso_events = defaultdict(int)

        # Wich values we must sum
        sum_values = {
            tso: [0] * len(table.sum_values)
            for tso in tso_frequency
            if tso is not None
        }

        # Must have sums
        sum_tso_events = defaultdict(int)
        year_sum_tso_events = {}

        # Run through every event
        for number in range(events_number):
            tso_name = self.events[titles.tso_name][number]
            sheet_name = self.get_short_name(tso_name)
            num_tso_events[tso_name] += 1

            # If tso first time
            if sheet_name not in workbook.sheetnames:
                # Create new sheet
                ws = workbook.create_sheet(title=sheet_name)
                # Insert main header
                ws = self.create_main_header(ws, tso_name, table)
                # Insert table header
                ws = self.create_header(ws, tables, titles, table)

            # Insert events
            ws = workbook[sheet_name]
            ws, sum_tso_events = self.create_content(ws,
                                                     tables,
                                                     titles,
                                                     table,
                                                     num_tso_events[tso_name],
                                                     number,
                                                     sum_tso_events,
                                                     year_sum_tso_events,
                                                     sum_values,
                                                     tso_name)

            # Insert footer
            if num_tso_events[tso_name] == tso_frequency[tso_name]:
                ws = self.create_footer(ws,
                                        num_tso_events[tso_name] + 1,
                                        tso_name,
                                        sum_tso_events,
                                        year_sum_tso_events,
                                        sum_values,
                                        tables,
                                        titles,
                                        table)
        return workbook

    def create_header(self,
                      ws,
                      tables,
                      titles,
                      table):

        # Set column number
        column = table.header_column

        # Set row number
        row = table.header_row

        # Set header style
        style = table.header_style

        # Set height row
        height = table.header_height

        # Where insert capex flow header
        capex_flow_pos = column + len(table.__dataclass_fields__) - 1

        # Create header
        for structure in table.__dataclass_fields__:
            structure = getattr(table, structure)

            if column == capex_flow_pos:
                # We must merge first row
                first_year = tables.terms[titles.first_year]
                last_year = tables.terms[titles.last_year]
                period = last_year - first_year + 1
                ws.merge_cells(
                    start_row=row, start_column=column,
                    end_row=row, end_column=column + period - 1
                )
                ws.cell(row=row, column=column).value = titles.total_cost

                # And draw head_capex_flow cells with dates in second row
                for i in range(period):
                    ws.cell(row=row, column=column + i).style = style
                for year in range(first_year, last_year + 1):
                    ws.cell(row=row + 1, column=column).value = year
                    ws.cell(row=row + 1, column=column).style = style
                    column += 1

            # Base header cell
            ws.merge_cells(
                start_row=row, start_column=column,
                end_row=row + 1, end_column=column
            )
            ws.cell(row=row, column=column).value = structure
            ws.cell(row=row, column=column).style = style
            ws.cell(row=row + 1, column=column).style = style
            width = table.get_width(structure)
            ws.column_dimensions[get_column_letter(column)].width = width
            column += 1

        ws.row_dimensions[row + 1].height = height
        return ws

    def create_footer(self,
                      ws,
                      serial_number,
                      tso_name,
                      sum_tso_events,
                      year_sum_tso_events,
                      sum_values,
                      tables,
                      titles,
                      table):

        # Set column number
        column = table.header_column

        # Set row number
        row = serial_number + table.header_row + 1

        # Set header style
        style = table.footer_style

        # Where insert capex flow footer
        capex_flow_pos = column + len(table.__dataclass_fields__) - 1
        for structure in table.__dataclass_fields__:
            structure = getattr(table, structure)

            if structure == table.number:
                # Number of order of this event in tso table
                ws.cell(row=row, column=column).value = serial_number

            elif column == 2:
                value = table.total
                ws.cell(row=row, column=column).value = value

            elif structure in table.sum_values:
                value = sum_values[tso_name][
                    table.sum_values[structure]
                ]
                ws.cell(row=row, column=column).value = (
                    value if value else '-'
                )

            elif structure == table.total_cost:
                # For this structure we need obtain capex
                value = sum_tso_events.pop(tso_name, 0)
                ws.cell(row=row, column=column).value = value

            elif structure == table.terms:
                value = self.get_sum_terms(year_sum_tso_events[tso_name],
                                           tables,
                                           titles)
                ws.cell(row=row, column=column).value = value

            elif column == capex_flow_pos:
                # For this structure we need obtain capex flow
                for value in year_sum_tso_events[tso_name]:
                    ws.cell(row=row, column=column).value = value
                    ws.cell(row=row, column=column).style = style
                    ws = self.set_number_format(ws,
                                                structure,
                                                titles,
                                                row,
                                                column)
                    column += 1

            if structure == titles.total:
                value = sum(year_sum_tso_events[tso_name])
                ws.cell(row=row, column=column).value = value
                del year_sum_tso_events[tso_name]

            ws.cell(row=row, column=column).style = style
            ws = self.set_number_format(ws,
                                        structure,
                                        titles,
                                        row,
                                        column)

            # Next column
            column += 1
        return ws


class SourceEvents(TableMixin, BaseMixin):

    def create_content(self,
                       ws,
                       tables,
                       titles,
                       table,
                       serial_number,
                       number,
                       sum_tso_events,
                       year_sum_tso_events,
                       sum_values,
                       tso_name):

        # Set style
        style = table.base_style

        # Set column number
        column = table.header_column

        # Set row number
        row = serial_number + table.header_row + 1

        # Where insert capex flow content
        capex_flow_pos = column + len(table.__dataclass_fields__) - 1
        # Create content
        for structure in table.__dataclass_fields__:
            structure = getattr(table, structure)
            if structure == table.number:
                # Number of order of this event in tso table
                ws.cell(row=row, column=column).value = serial_number

            elif structure == table.total_cost:

                # For this structure we need obtain capex
                unit_type = self.get_unit_type(titles, number)
                power, is_tfu = self.get_power(table, number)
                capex = tables.get_energy_source_capex(
                    power, unit_type, is_tfu, titles
                )
                ws.cell(row=row, column=column).value = capex
                sum_tso_events[tso_name] += capex

            # Save sum powers for footer
            elif structure in table.sum_values:
                value = self.events[structure][number]
                sum_values[tso_name][
                    table.sum_values[structure]
                ] += 0 if value == '-' else value
                ws.cell(row=row, column=column).value = value

            elif column == capex_flow_pos:
                # For this structure we need obtain capex flow
                terms = self.events[titles.terms][number]
                capex_flow = tables.get_capex_flow(capex,
                                                   terms,
                                                   titles.energy_sources,
                                                   titles)
                if serial_number == 1:
                    year_sum_tso_events[tso_name] = capex_flow

                for index, value in enumerate(capex_flow):
                    ws.cell(row=row, column=column).value = value
                    ws.cell(row=row, column=column).style = style
                    ws = self.set_number_format(ws,
                                                structure,
                                                titles,
                                                row,
                                                column)
                    if serial_number > 1:
                        year_sum_tso_events[tso_name][index] += (
                            capex_flow[index]
                        )
                    column += 1

            else:
                # This structures goes unchanged
                value = self.events[structure][number]
                ws.cell(row=row, column=column).value = value

            if structure == titles.total:
                # For this structure we need obtain sum of capex flow
                ws.cell(row=row, column=column).value = sum(capex_flow)

            ws.cell(row=row, column=column).style = style
            ws = self.set_number_format(ws, structure, titles, row, column)
            # Next column
            column += 1
        return ws, sum_tso_events

    def get_power(self,
                  table,
                  number):
        mw = self.events[table.mw][number]
        th = self.events[table.th][number]
        gh = self.events[table.gh][number]

        if mw != '-':
            power = mw
            index = table.sum_values[table.mw]
        elif th != '-':
            power = 0.6 * th
            index = table.sum_values[table.th]
        else:
            power = gh
            index = (
                table.sum_values[table.gh]
            )
        return power, index
