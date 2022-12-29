from openpyxl import load_workbook


class OriginTables:
    def __init__(self, titles) -> None:

        wb = load_workbook(filename=titles.filename, data_only=True)
        table_names = {}
        for worksheet in wb.sheetnames:
            for table in wb[worksheet].tables:
                table_names[table] = worksheet

        self.energy_source_unit_costs = (
            self.__get_data(wb,
                            table_names,
                            titles.energy_source_unit_costs)
        )
        self.heating_network_unit_costs = (
            self.__get_data(wb,
                            table_names,
                            titles.heating_network_unit_costs)
        )
        self.tfu_unit_cost = (
            self.__get_data(wb, table_names, titles.tfu_unit_cost)
        )
        self.chp_unit_costs = (
            self.__get_data(wb, table_names, titles.chp_unit_costs)
        )
        self.deflators = self.__get_data(wb, table_names, titles.deflators)
        self.terms = self.__get_data_from_row(wb, table_names, titles.terms)
        self.stages = self.__get_data_from_header(wb,
                                                  table_names,
                                                  titles.stages)
        self.nds = self.__get_data_from_row(wb, table_names, titles.nds)
        self.tso_list = self.__get_data_from_col(wb,
                                                 table_names,
                                                 titles.tso_list)
        self.energy_source_events = (
            self.__get_data(wb, table_names, titles.energy_source_events)
        )
        self.heating_network_events = (
            self.__get_data(wb, table_names, titles.heating_network_events)
        )

    def __data_in_dataclasses(self):
        pass

    def __get_data(self, wb, names, name) -> dict:
        '''Collect datas as shown below
        dict = {
            header_1: column_1_datas,
            header_2: column_2_datas,
            ...
        }
        '''
        worksheet = wb[names[name]]
        table_range = worksheet.tables[name].ref

        table_head = worksheet[table_range][0]
        table_data = worksheet[table_range][1:]

        columns = [column.value for column in table_head]
        data = {column: [] for column in columns}

        for row in table_data:
            row_value = [cell.value for cell in row]
            for key, value in zip(columns, row_value):
                data[key].append(value)

        return data

    def __get_data_from_header(self, wb, names, name) -> dict:
        '''Collect datas as shown below
        It is assumed that the first column is always a serial number
        dict = {
            data(row_2, column_2):
                {
                    header_3: data(row_2, column_3),
                    header_4: data(row_2, column_3),
                    ...
                },
            data(row_3, column_2):
                {
                    header_3: data(row_3, column_3),
                    header_4: data(row_3, column_3),
                    ...
                },
            ...
        }
        '''
        worksheet = wb[names[name]]
        table_range = worksheet.tables[name].ref

        table_head = worksheet[table_range][0]
        table_data = worksheet[table_range][1:]

        data = {}
        for row in table_data:
            for index, cell in enumerate(row):
                if index == 1:
                    data[cell.value] = {}
                elif index > 1:
                    data[row[1].value][table_head[index].value] = cell.value
        return data

    def __get_data_from_col(self, wb, names, name) -> dict:
        '''Collect datas as shown below
        It is assumed that the first column is always a serial number
        It is assumed that the table has only two rows include header
        dict = {
            data(row_2, column_2): data(row_2, column_3),
            data(row_3, column_2): data(row_3, column_3),
            data(row_3, column_2): data(row_4, column_3),
            ...
        }
        '''
        worksheet = wb[names[name]]
        table_range = worksheet.tables[name].ref

        table_data = worksheet[table_range][1:]
        data = {}
        for row in table_data:
            for index, cell in enumerate(row):
                data[row[1].value] = row[2].value
        return data

    def __get_data_from_row(self, wb, names, name) -> dict:
        '''Collect datas as shown below
        It is assumed that the first column is always a serial number
        It is assumed that the table has only two rows include header
        dict = {
            header_1: data(row_2, column_1),
            header_2: data(row_2, column_2),
            header_3: data(row_2, column_3),
            ...
        }
        '''
        worksheet = wb[names[name]]
        table_range = worksheet.tables[name].ref

        table_head = worksheet[table_range][0]
        table_data = worksheet[table_range][1]
        data = {}
        for index, column in enumerate(table_head):
            data[table_head[index].value] = table_data[index].value
        return data

    def __binary_search(self, value: float, array: list) -> int:
        right = len(array) - 1
        left = 0
        while right - left != 1:
            center = (left + right) // 2
            if array[center] == value:
                return center
            elif value > array[center]:
                left = center
            else:
                right = center
        return left if array[right] - value > value - array[left] else right

    def get_energy_source_capex(self,
                                power: float,
                                unit_type: str,
                                title: str,
                                is_chp=False,
                                is_tfu=False,
                                ) -> float:

        if is_tfu:
            unit_cost = self.tfu_unit_cost.get(unit_type)[0]
        else:
            if is_chp:
                powers = self.chp_unit_costs.get(title)
                unit_costs = self.chp_unit_costs.get(unit_type)
            else:
                powers = self.energy_source_unit_costs.get(title)
                unit_costs = self.energy_source_unit_costs.get(unit_type)
            unit_cost = unit_costs[self.__binary_search(power, powers)]

        return power * unit_cost

    def get_heating_network_capex(self,
                                  diameter: float,
                                  length: float,
                                  laying_type: str,
                                  unit_type: str) -> float:
        diameters = self.heating_network_unit_costs.get('2Ду, мм')
        unit_type = '' if unit_type == 'строительство' else '2'
        unit_costs = self.heating_network_unit_costs.get(laying_type
                                                         + unit_type)
        # print(f'{diameter=}, {diameters=}')
        # print(unit_costs)
        unit_cost = unit_costs[self.__binary_search(diameter, diameters)]
        return length * unit_cost / 1000

    def get_capex_flow(self,
                       capex: float,
                       time: str,
                       otype: str,
                       titles) -> dict:
        time = str(time)
        if time[:4] == time[-4:]:
            start = end = int(time)
        else:
            start = int(time[:4])
            end = int(time[-4:])
        # create capex flow
        capex_flow = []
        time = end - start + 1
        deflator = 1
        design_rate = self.stages[titles.design][otype] / 100
        for index, year in enumerate(self.deflators[titles.year]):
            if year >= self.terms[titles.year_cost]:
                deflator *= self.deflators[titles.deflator][index]

            # fill the capex flow
            if (
                self.terms[titles.first_year] <= year
                <= self.terms[titles.last_year]
            ):
                if start <= year <= end:
                    if time == 1:
                        capex_flow.append(capex * deflator)
                    elif year - start:
                        capex_flow.append(
                            capex * (1 - design_rate) * deflator / (time - 1)
                        )
                    else:
                        # in the first year we carry out design and survey work
                        capex_flow.append(capex * design_rate * deflator)
                    capex_flow[-1] *= (1 + self.nds[titles.nds])
                else:
                    capex_flow.append(0)

        return capex_flow
